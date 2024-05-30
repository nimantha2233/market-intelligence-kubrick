import pandas as pd
import os
import yfinance as yf
import json
import datetime
import openpyxl
from datetime import datetime, date

def get_company_info_pricereport(company_name, ticker):
    """
    Produces price report for a given company
    
    Args:
    company_name (str): Full company name
    ticker (str): Company ticker. This can be an empty string if the company is private.

    Return:
    (json) : Financial Data to produce the price report
    """
    file_path = "Kubrick MI Data.xlsx"
    sheet_name = "Price Report"

    if ticker == "":
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "Currency": "Private",
            "52 Week Range": "Private"
        }
    
    try:
        # Fetch company data from Yahoo Finance
        stock = yf.Ticker(ticker)
        info = stock.info
        # Check if the returned info indicates data not found
        if len(info) <= 1:
            print(f"Issue with with ticker: {company_name} : {info}")
            return {
                "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
                "Company Name": company_name,
                "Previous Close": "Private",
                "%-Change": "Private",
                "Sector": "Private",
                "Currency": "Private",
                "52 Week Range": "Private"
            }
        # Try to get historical data
        previous_close = info.get("previousClose", "N/A")
        if previous_close == "N/A":
            print(f"Error while retrieving financial data for: {ticker}.")
            return {
                "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
                "Company Name": company_name,
                "Previous Close": "Private",
                "%-Change": "Private",
                "Sector": "Private",
                "Currency": "Private",
                "52 Week Range": "Private"
            }
        

        # Open the Excel file and check for previous close
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            last_close = None
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                if row[1].value == company_name and isinstance(row[2].value, (int, float)):
                    last_close = row[2].value
                    break
            if last_close is None:
                percent_change = 0
            else:
                percent_change = abs(round(((previous_close - last_close) / last_close) * 100, 4))
        else:
            percent_change = 0
        wb.close()
        print(f"Successful save for ticker: {company_name}")
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": previous_close,
            "%-Change": percent_change,
            "Sector": info.get("sector", "No Sector Available"),
            "Currency": info.get('currency', "No Currency Available"),
            "52 Week Range": f'{info.get("fiftyTwoWeekLow", "No 52 Week Low Available")} - {info.get("fiftyTwoWeekHigh", "No 52 Week High Available")}'
        }
    except:
        print(f"Error while retrieving financial data for: {company_name}.")
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "Currency": "Private",
            "52 Week Range": "Private"
        }

def update_excel(data, file_path = "Kubrick MI Data.xlsx"):

    """
    Updates the sheet "Price Report" with the financial data from a company
    
    Args:
    data (json) : Financial Company data
    """
    sheet_name = "Price Report"

    try:
        # Load workbook and check for sheet
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            # Add headers
            headers = list(data.keys())
            ws.append(headers)
        else:
            ws = wb[sheet_name]
        
        # Find if the company already exists and remove the row if it does
        company_col = 2  # Assuming "Company Name" is in column B (index 1)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            if row[company_col - 1].value == data["Company Name"]:
                ws.delete_rows(row[0].row)
                break
        
        # Append new data
        ws.append(list(data.values()))

        # Save workbook
        wb.save(file_path)
    except Exception as e:
        print(f"Error updating Excel file: {e}")

temp_dict = {
    'Capgemini SE' : 'CAP.PA',
    'AND Digital' : '',
    'Dufrain' : '',
    'FDM Group (Holdings) Ltd' : 'FDM.L',
    'Slalom' : '',
    'Tata Consultancy Services Limited' : 'TCS.BO',
    'Wipro Limited' : 'WIT',
    'Fjord Consulting Group' : '',
    'Mesh AI' : '',
    'Sparta Global' : '',
    'Ten10' : '',
    'Credera' : '',
    'Infinite Lambda' : '',
    'BetterGov' : '',
    'Cambridge Consultants Limited' : '',
    'Capco Limited' : '',
    'Cognizant Technology Solutions Corporation' : 'CTSH',
    'Infosys Limited' : 'INFY',
    'IQVIA Holdings Inc' : 'IQV',
    'Kubrick Group Limited' : '',
    '11:FS' : '',
    'a1qa' : '',
    'Accenture PLC' : 'ACN',
    'Adatis' : '',
    'Afiniti' : '',
    'Airwalk Reply' : '',
    'Alchemmy' : '',
    'Alix Partners' : '',
    'Apexon' : '',
    'Arthur D Little' : '',
    'Atos Group' : '', # Unsure on what the ticker is? Mentioned it is public in excel by Simon
    'Atos SE' : 'ATO.PA',
    'Avanade Inc.' : '',
    'Bain & Company' : '',
    'Baringa' : '',
    'Billigence' : '',
    'BJSS' : '',
    'Blueberry Consultants Ltd.' : '',
    'Booz Allen Hamilton Holding Corp' : 'BAH',
    'Boston Consulting Group' : '',
    'Broadstones Tech' : '',
    'Cambridge Design Partnership' : '',
    'Canon Inc.' : '7751.T',
    'Capita' : 'CPI.L',
    'Centric Consulting' : '',
    'CGI' : '',
    'CIGNITI Technologies Ltd' : 'CIGNITITEC.BO',
    'Clarasys' : '',
    'Cognifide' : '',
    'Contino Ltd.' : '',
    'Credo Consulting' : '',
    'Deloitte' : '',
    'Designit' : '',
    'Digital Workplace Group' : '',
    'DXC Technology' : 'DXC',
    'Eclature Technologies' : '',
    'Eden McCallum' : '',
    'Edge Testing Solutions' : '',
    'Elixirr International Plc' : 'ELIX.L',
    'EPAM Systems Inc' : 'E3M.BE',
    'Equal Experts' : '',
    'EY' : '',
    'Faculty AI' : '',
    'Frog Design' : '',
    'Frontier Economics' : '',
    'GeekTek' : '',
    'Genpact Ltd' : '35G.SG',
    'HCL Technologies Ltd' : 'HCLTECH.BO',
    'Hexaware Technologies Ltd' : '',
    'HP Inc' : 'HPQ',
    'Icon PLC' : 'IJF.SG',
    'IDEO' : '',
    'Infostretch Corp' : '',
    'International Business Machines Corp' : 'IBM',
    'JMAN Group' : '',
    'Kainos' : 'KNOS.L',
    'Kearney' : '',
    'KONICA MINOLTA INC' : 'KNCAF',
    'KPMG' : '',
    'Laboratory Corp Of America Holdings' : 'LAB.F',
    'LockPath, Inc.' : '',
    'LogicManager' : '',
    'LogicSource, Inc.' : '',
    'Lovelytics' : '',
    'Lunar' : '',
    'Made Tech Group Plc ' : 'MTEC.L',
    'Mason Advisory' : '',
    'McKinsey' : '',
    'METRICSTREAM INC' : '',
    'MindTree Ltd' : '',
    'Mosaic Island' : '',
    'Mphasis Ltd' : 'MPHASIS.BO',
    'mthree' : '',
    'Neoris' : '',
    'NexInfo' : '',
    'North Highland' : '',
    'OC&C Strategy Consultants' : '',
    'Oliver Wyman' : '',
    'OpenCredo' : '',
    'Oracle Consulting' : '',
    'PA Consulting Group' : '',
    'Parexel International Corp' : '',
    'Peru Consulting' : '',
    'Planit Testing' : '',
    'PPD Inc' : '',
    'PRA Health Sciences Inc' : '',
    'Project One' : '',
    'Projective' : '',
    'Prolifics' : '',
    'PwC' : 'PWC.F',
    'RICOH CO LTD' : 'RICO.L',
    'Roland Berger' : '',
    'SEIKO EPSON CORP' : 'SE7.DU',
    'Simon Kutcher & Partners' : '',
    'SkillStorm' : '',
    'Softwire' : '',
    'Spring Studios' : '',
    'Strategy & London' : '',
    'Sutherland' : '',
    'Switchfast Technologies' : '',
    'Syneos Health Inc' : '',
    'TAG Solutions' : '',
    'Tech Mahindra Ltd' : 'TECHM.BO',
    'Terillium' : '',
    'Testhouse Ltd' : '',
    'The Berkeley Partnership' : '',
    'The PSC' : '',
    'Thoughtworks' : '7W8.F',
    'Ultimus Fund Solutions' : '',
    'Verisk Analytics Inc' : 'VA7A.SG',
    'Wavestone' : '0G1T.L',
    'Xerox Corp' : '',
    'Zoonou' : '',
    'ZS' : '',
}

# for yfinance_name, yfinance_ticker in temp_dict.items():
#     temp_data = get_company_info_pricereport(yfinance_name, yfinance_ticker)
#     update_excel(temp_data)

import csv

def save_companies_with_tickers(temp_dict, filename='companies_with_tickers.csv'):
    # Filter out companies without tickers
    filtered_dict = {company: ticker for company, ticker in temp_dict.items() if ticker}
    
    # Write to CSV
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        # Write header
        writer.writerow(['company_name', 'ticker'])
        # Write company names and tickers
        for company, ticker in filtered_dict.items():
            writer.writerow([company, ticker])

save_companies_with_tickers(temp_dict)