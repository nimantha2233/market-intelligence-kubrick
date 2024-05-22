import yfinance as yf
from datetime import datetime
import openpyxl

def get_company_info_pricereport(company_name, ticker):

    file_path = "Kubrick MI Data.xlsx"
    sheet_name = "Price Report"

    if ticker == "":
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "52 Week Range": "Private"
        }
    
    try:
        # Fetch company data from Yahoo Finance
        stock = yf.Ticker(ticker)
        info = stock.info
        # Check if the returned info indicates data not found
        if 'trailingPegRatio' in info and info['trailingPegRatio'] is None:
            return {
                "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
                "Company Name": company_name,
                "Previous Close": "Private",
                "%-Change": "Private",
                "Sector": "Private",
                "52 Week Range": "Private"
            }
        # Try to get historical data
        hist = stock.history(period="1mo")
        
        if not hist.empty:
            last_close = hist['Close'].iloc[-2]
            prev_close = hist['Close'].iloc[-1]
            percent_change = round(((prev_close - last_close) / last_close) * 100, 4)
        else:
            prev_close = info.get("previousClose", "N/A")
            # Open the Excel file and check for previous close
            wb = openpyxl.load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                last_close = None
                for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                    if row[1].value == company_name and isinstance(row[2].value, (int, float)):
                        last_close = row[2].value
                        break
                if last_close is None:
                    percent_change = 0
                else:
                    percent_change = round(((prev_close - last_close) / last_close) * 100, 4)
            else:
                percent_change = 0
            wb.close()

        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": info.get("previousClose", "N/A"),
            "%-Change": percent_change,
            "Sector": info.get("sector", "N/A"),
            "52 Week Range": f'{info.get("fiftyTwoWeekLow", "N/A")} - {info.get("fiftyTwoWeekHigh", "N/A")}'
        }
    except:
        print(f"Error while retrieving financial data for: {ticker}.")
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "52 Week Range": "Private"
        }

def update_excel(data):

    file_path = "Kubrick MI Data.xlsx"
    sheet_name = "Price Report"

    try:
        # Load workbook and check for sheet
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            # Add headers
            headers = list(data.keys())
            ws.append(headers)
        else:
            ws = wb[sheet_name]
        
        # Find if the company already exists and remove the row if it does
        company_col = 2  # Assuming "Company Name" is in column B (index 1)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            if row[company_col - 1].value == data["Company Name"]:
                ws.delete_rows(row[0].row)
                break
        
        # Append new data
        ws.append(list(data.values()))

        # Save workbook
        wb.save(file_path)
    except Exception as e:
        print(f"Error updating Excel file: {e}")

ticker = "BetterGov"
company_name = "BetterGov"

company_info = get_company_info_pricereport(company_name, ticker)
if company_info:
    update_excel(company_info)

