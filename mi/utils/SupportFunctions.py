import pandas as pd
import os
import yfinance as yf
import json
import datetime
import openpyxl
from datetime import datetime, date

def sheet_exists(file_path, sheet_name):
    """
    Checks if a sheet exists within xlsx file
    
    Args:
    file_path (str): Location of xlsx file
    sheet_name (str): The name of the sheet to update or delete.

    Returns:
    Boolean : True if sheet exists, false if it doesn't.
    """
    if os.path.exists(file_path):
        xl = pd.ExcelFile(file_path)
        return sheet_name in xl.sheet_names
    else:
        return False

def write_to_excel(df, file_path, sheet_name):
    """
    Saves the df in the sheet in the excel workbook.
    
    Args:
    df (pd.DataFrame): Dataframe to be stored
    file_path (str): Location of xlsx file
    sheet_name (str): The name of the sheet to update or delete.

    Returns:

    """
    if sheet_exists(file_path, sheet_name):
        # If the sheet exists, overwrite all the data in that sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
            workbook.save(file_path)  # Save the workbook without the sheet
        
        # Reopen the workbook with pd.ExcelWriter and add the new sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If the sheet doesn't exist, append a new sheet with the data
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def read_from_excel(file_path, sheet_name):
    """
    Reads a specific sheet from the excel workbook and
    returns it as a dataframe
    
    Args:
    file_path (str): Location of xlsx file
    sheet_name (str): The name of the sheet to update or delete.

    Returns:
    df (pd.DataFrame): Dataframe from the sheet in excel. If the sheet doesn't exist, returns an empty DataFrame.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except ValueError:
        return read_template()
    return df

def read_template():
    """
    Imports the template table from Template.xlsx
    
    Args:

    Returns:
    df (pd.DataFrame): Template table as dataframe
    """
    try:
        # Read the Excel file into a DataFrame
        df = pd.read_excel(r"C:\Users\NimanthaFernando\Innovation_Team_Projects\Market_Intelligence\MI\mi\utils\Template.xlsx", sheet_name='Template')
        return df
    except FileNotFoundError:
        print("Template.xlsx not found. Make sure the file exists in the current directory.")

def get_company_status(symbol):
    """
    Provides information on whether the company is private or public
    
    Args:
    symbol (str) : Company symbol to obtain yfinance data. This could be blank if company doesn't have a symbol

    Returns:
    (str) : Either Private or Public
    """
    if symbol == "":
        return "Private"
    try:
        stock = yf.Ticker(symbol)
        info = stock.info
        if info['quoteType'] == 'EQUITY':
            return "Public"
        else:
            return "Private"
    except KeyError:
        return 'Private'
    except Exception as e:
        return str(e)

def get_company_details(symbol):
    """
    Obtains yfinance data if company is public.
    
    Args:
    symbol (str) : Company symbol to obtain yfinance data. This could be blank if company doesn't have a symbol

    Returns:
    (json) : Company information
    """


    status = get_company_status(symbol)
    if status == "Private":
        return json.dumps({"Public/Private": "Private"})
    elif status == "Public":
        try:
            stock = yf.Ticker(symbol)
            info = stock.info
            company_details = {
                "Public/Private": "Public",
                "Previous Close": info.get("previousClose", "N/A"),
                "52 Week Range": info.get("fiftyTwoWeekRange", "N/A"),
                "Sector": info.get("sector", "N/A"),
                "Industry": info.get("industry", "N/A"),
                "Full Time Employees": info.get("fullTimeEmployees", "N/A"),
                "Market Cap": info.get("marketCap", "N/A"),
                "Fiscal Year Ends": info.get("fiscalYearEnd", "N/A"),
                "Revenue": info.get("totalRevenue", "N/A"),
                "EBITDA": info.get("ebitda", "N/A"),
            }
            return json.dumps(company_details)
        except Exception as e:
            return json.dumps({"error": str(e)})
    else:
        return json.dumps({"error": status})

def create_final_df(company_name, url, json_data, df):
    """
    Creates the final dataframe to be stored using data from webscrape + financial data
    
    Args:
    company_name (str) : Name of company
    url (str) : URL of company
    json_data (json) : Financial data of company
    df (pd.Dataframe) : Webscrape data of comapny

    Returns:
    final_df (pd.Dataframe) : Dataframe with all data in the Template table format
    """
    # Define column list
    columns_list = read_template().columns.tolist()

    # Load JSON data
    data = json.loads(json_data)

    # Initialize an empty list to store dictionaries representing rows
    rows = []

    # Get the current date
    date_of_collection = str(date.today())

    # Iterate over each row in the expertise dataframe
    for index, row in df.iterrows():
        # Create a dictionary to store data for the current row
        row_data = {'Date of Collection': date_of_collection,
                    'Company Name': company_name,
                    'Website_URL': url}

        # Add other data from the row
        row_data['Practices'] = row['Practices']
        row_data['Practices_URL'] = row['Practices_URL']

        # Check if 'Services' key exists in the row
        if 'Services' in row:
            row_data['Services'] = row['Services']
        else:
            row_data['Services'] = " "  # or any default value you prefer

        # Check if 'Services_URL' key exists in the row
        if 'Services_URL' in row:
            row_data['Services_URL'] = row['Services_URL']
        else:
            row_data['Services_URL'] = " "  # or any default value you prefer

        # Check if 'Solutions' key exists in the row
        if 'Solutions' in row:
            row_data['Solutions'] = row['Solutions']
        else:
            row_data['Solutions'] = " "  # or any default value you prefer
        
                # Check if 'Solutions_URL' key exists in the row
        if 'Solutions_URL' in row:
            row_data['Solutions_URL'] = row['Solutions_URL']
        else:
            row_data['Solutions_URL'] = " "  # or any default value you prefer


        # Add data from JSON
        for key, value in data.items():
            row_data[key] = value

        # Add other columns from columns_list
        for col in columns_list:
            if col not in row_data:
                row_data[col] = " "

        # Append the row_data to the list of rows
        rows.append(row_data)

    # Create the final dataframe from the list of rows
    final_df = pd.DataFrame(rows)

    # Reorder the columns according to column_names list
    final_df = final_df[columns_list]

    return final_df

def log_differences(action, sheet_name, num_differences):
    """
    Log the differences found to a text file.
    
    Args:
    action (str): The action taken (e.g., "Added" or "No differences found").
    sheet_name (str): The name of the sheet.
    num_differences (int): The number of differences found.
    """
    if num_differences == 0:
        log_message = f"{datetime.now().timestamp()}: No new data for company: '{sheet_name}'"
    elif num_differences == 1:
        log_message = f"{datetime.now().timestamp()}: {action} {num_differences} new data row for company: '{sheet_name}'"
    else:
        log_message = f"{datetime.now().timestamp()}: {action} {num_differences} new data rows for company: '{sheet_name}'"
    with open('log_diff.txt', 'a') as log_file:
        log_file.write(log_message + '\n')

def log_new_and_modified_rows(df1, df2, sheet_name):
    """
    Find rows that are new or modified in df1 compared to df2 and handle Excel file updates.
    
    Args:
    df1 (pd.DataFrame): The most up-to-date dataframe.
    df2 (pd.DataFrame): The older dataframe to compare against.
    sheet_name (str): The name of the sheet to update or delete.
    """
    # Convert columns to the same data type
    # df1['Date of Collection'] = pd.to_datetime(df1['Date of Collection'])
    # df2['Date of Collection'] = pd.to_datetime(df2['Date of Collection'])


    # First, we reset the index to ensure the comparison is done row-wise
    df1_reset = df1.reset_index(drop=True)
    df2_reset = df2.reset_index(drop=True)
    df1_reset.to_csv(r'./df1.csv')
    df2_reset.to_csv(r'./df2.csv')

    excel_file = "Kubrick MI Diff.xlsx"

    # Merge to find new and modified rows
    merged_df = df1_reset.merge(df2_reset, how='left', indicator=True)

    # Select new rows
    new_rows = merged_df[merged_df['_merge'] == 'left_only'].drop(columns='_merge')
    # new_rows.to_csv(r'./new_rows.csv') # empty

    # Select potentially modified rows (those that are in both dataframes)
    common_rows = merged_df[merged_df['_merge'] == 'both'].drop(columns='_merge')

    common_rows.to_csv(r'./common_rows.csv')

    # Identify rows that have changes by comparing the content
    modified_mask = (df1_reset.loc[common_rows.index] != df2_reset.loc[common_rows.index]).any(axis=1)
    modified_rows = df1_reset.loc[common_rows.index][modified_mask]


    
    # Combine new and modified rows
    new_and_modified_df = pd.concat([new_rows, modified_rows])

    if not os.path.exists(excel_file):
        # Create an empty Excel file with just the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet 1", index=False)

    if not new_and_modified_df.empty:
        # If there are new or modified rows, write them to the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            new_and_modified_df.to_excel(writer, sheet_name=sheet_name, index=False)
        log_differences("Added", sheet_name, len(new_and_modified_df))
    else:
        # If there are no new or modified rows, remove the sheet if it exists
        try:
            workbook = openpyxl.load_workbook(excel_file)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
                workbook.save(excel_file)
        except FileNotFoundError:
            # If the file doesn't exist, do nothing
            pass
        log_differences("No differences found", sheet_name, 0)


def remove_duplicates(soup_list) -> list:
    """
    Find duplicate soup objects and remove them 
    
    Args:
    soup_list (list): list of bs4.element.ResultSet objects 
    """
    unique_soups = []
    unique_strings = set()
    for soup in soup_list:
        soup_str = str(soup)
        if soup_str not in unique_strings:
            unique_soups.append(soup)
            unique_strings.add(soup_str)
    return unique_soups


