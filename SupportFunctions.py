import pandas as pd
import os
import yfinance as yf
import json
import datetime
import openpyxl
from datetime import datetime, date
import scrapers
import numpy as np

def sheet_exists(file_path, sheet_name):
    """
    Checks if a sheet exists within xlsx file
    
    Args:
    file_path (str): Location of xlsx file
    sheet_name (str): The name of the sheet to update or delete.

    Returns:
    Boolean : True if sheet exists, false if it doesn't.
    """
    if os.path.exists(file_path):
        xl = pd.ExcelFile(file_path)
        return sheet_name in xl.sheet_names
    else:
        return False

def company_intel_table(company_name, url, file_path="kubrick_mi_company_intel.csv"):

    if not os.path.exists(file_path):
        empty_df = pd.DataFrame(columns=['Date Collected', 'Company Name', 'URL','Training Program Duration',
                                         'Anecdotal Views of Quality', 'Consultant Pricing'])  # Assuming these are your column names
        empty_df.to_csv(file_path, index=False)

    try:
        data = pd.read_csv(file_path, encoding='utf-8')
    except UnicodeDecodeError:
        data = pd.read_csv(file_path, encoding='ISO-8859-1')

    data_as_dict = defaultdict(list)

    company_names_column = data.iloc[:, 1]
    company_exists = company_name in company_names_column.values

    data_as_dict["Date Collected"].append(datetime.now().strftime("%Y-%m-%d"))
    data_as_dict["Company Name"].append(company_name)
    data_as_dict["URL"].append(url)

    training_duration_index = data.columns.get_loc("Training Program Duration")
    try:
        if company_exists:
            company_index = company_names_column[company_names_column == company_name].index[0]
            for col in data.columns[training_duration_index:]:
                data_as_dict[col].append(data.at[company_index, col])
            
            for key, value in data_as_dict.items():
                data.at[company_index, key] = value[0]
            
        else:
            for col in data.columns[training_duration_index:]:
                if col not in data_as_dict:
                    data_as_dict[col].append("To be completed manually")
                new_data = pd.DataFrame(data_as_dict)
                data = pd.concat([data, new_data], ignore_index=True)
    except Exception as e:
        log_error(f"Error obtaining intel data for {company_name}: {e}")
        return False
    
    for column in data.columns[3:]:
        data[column] = [f'No {column} Available' if item in ('', ' ', None, 'nan', np.nan) else item for item in data[column]]
    data.dropna(inplace=True)
    data.reset_index(drop=True, inplace=True)
    data.to_csv(file_path, index=False, encoding='ISO-8859-1')
    return True

def write_to_excel(df, file_path, sheet_name):
    """
    Saves the df in the sheet in the excel workbook.
    
    Args:
    df (pd.DataFrame): Dataframe to be stored
    file_path (str): Location of xlsx file
    sheet_name (str): The name of the sheet to update or delete.

    Returns:

    """
    if sheet_exists(file_path, sheet_name):
        # If the sheet exists, overwrite all the data in that sheet
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
            workbook.save(file_path)  # Save the workbook without the sheet
        
        # Reopen the workbook with pd.ExcelWriter and add the new sheet
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # If the sheet doesn't exist, append a new sheet with the data
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def read_from_excel(file_path, sheet_name, template_file_path):
    """
    Reads a specific sheet from the excel workbook and
    returns it as a dataframe
    
    Args:
    file_path (str): Location of xlsx file
    sheet_name (str): The name of the sheet to update or delete.

    Returns:
    df (pd.DataFrame): Dataframe from the sheet in excel. If the sheet doesn't exist, returns an empty DataFrame.
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df = table_sorter(df)
    except ValueError:
        return read_template(template_file_path)
    return df

def read_template(template_file_path):
    """
    Imports the template table from Template.xlsx
    
    Args:

    Returns:
    df (pd.DataFrame): Template table as dataframe
    """
    try:
        # Read the Excel file into a DataFrame
        df = pd.read_excel(template_file_path, sheet_name='Template')
        return df
    except FileNotFoundError:
        print("Template.xlsx not found. Make sure the file exists in the current directory.")

def get_company_status(symbol):
    """
    Provides information on whether the company is private or public
    
    Args:
    symbol (str) : Company symbol to obtain yfinance data. This could be blank if company doesn't have a symbol

    Returns:
    (str) : Either Private or Public
    """
    if symbol == "":
        return "Private"
    try:
        stock = yf.Ticker(symbol)
        info = stock.info
        if info['quoteType'] == 'EQUITY':
            return "Public"
        else:
            return "Private"
    except KeyError:
        return 'Private'
    except Exception as e:
        return str(e)

def get_company_details(symbol):
    """
    Obtains yfinance data if company is public.
    
    Args:
    symbol (str) : Company symbol to obtain yfinance data. This could be blank if company doesn't have a symbol

    Returns:
    (json) : Company information
    """


    status = get_company_status(symbol)
    if status == "Private":
        company_details = {
                "Previous Close": "Private",
                "52 Week Range": "Private",
                "Sector": "Private",
                "Industry": "Private",
                "Full Time Employees": "Private",
                "Market Cap": "Private",
                "Fiscal Year Ends": "Private",
                "Revenue": "Private",
                "EBITDA": "Private",
        }
        return json.dumps(company_details)
    elif status == "Public":
        try:
            stock = yf.Ticker(symbol)
            info = stock.info
            company_details = {
                "Previous Close": info.get("previousClose", np.nan),
                "52 Week Range": f'{info.get("fiftyTwoWeekLow", np.nan)} - {info.get("fiftyTwoWeekHigh", np.nan)}',
                "Sector": info.get("sector", np.nan),
                "Industry": info.get("industry", np.nan),
                "Full Time Employees": info.get("fullTimeEmployees", np.nan),
                "Market Cap": info.get("marketCap", np.nan),
                "Fiscal Year Ends": info.get("fiscalYearEnd", np.nan),
                "Revenue": info.get("totalRevenue", np.nan),
                "EBITDA": info.get("ebitda", np.nan),
            }
            return json.dumps(company_details)
        except Exception as e:
            return json.dumps({"error": str(e)})
            #return f'Error, check {symbol} is a valid ticker'
    else:
        return json.dumps({"error": status})
    #return company_details
    

def create_final_df(company_name, url, json_data, df, template_file_path):
    """
    Creates the final dataframe to be stored using data from webscrape + financial data
    
    Args:
    company_name (str) : Name of company
    url (str) : URL of company
    json_data (json) : Financial data of company
    df (pd.Dataframe) : Webscrape data of comapny

    Returns:
    final_df (pd.Dataframe) : Dataframe with all data in the Template table format
    """
    # Define column list
    columns_list = read_template(template_file_path).columns.tolist()

    # Load JSON data
    data = json.loads(json_data)

    # Initialize an empty list to store dictionaries representing rows
    rows = []

    # Get the current date
    date_of_collection = date.today()

    # Iterate over each row in the expertise dataframe
    for index, row in df.iterrows():
        # Create a dictionary to store data for the current row
        row_data = {'Date of Collection': date_of_collection,
                    'Company Name': company_name,
                    'Website_URL': url}

        # Add other data from the row
        row_data['Practices'] = row['Practices']
        row_data['Practices_URL'] = row['Practices_URL']

        # Check if 'Services' key exists in the row
        if 'Services' in row:
            row_data['Services'] = row['Services']
        else:
            row_data['Services'] = "No Services"  # or any default value you prefer

        # Check if 'Services_URL' key exists in the row
        if 'Services_URL' in row:
            row_data['Services_URL'] = row['Services_URL']
        else:
            row_data['Services_URL'] = "No Services URL"  # or any default value you prefer

        # Check if 'Solutions' key exists in the row
        if 'Solutions' in row:
            row_data['Solutions'] = row['Solutions']
        else:
            row_data['Solutions'] = "No Solutions"  # or any default value you prefer
        
                # Check if 'Solutions_URL' key exists in the row
        if 'Solutions_URL' in row:
            row_data['Solutions_URL'] = row['Solutions_URL']
        else:
            row_data['Solutions_URL'] = "No Solutions URL"  # or any default value you prefer


        # Add data from JSON
        for key, value in data.items():
            row_data[key] = value

        # Add other columns from columns_list
        for col in columns_list:
            if col not in row_data:
                row_data[col] = " "

        # Append the row_data to the list of rows
        rows.append(row_data)

    # Create the final dataframe from the list of rows
    final_df = pd.DataFrame(rows)

    # Reorder the columns according to column_names list
    final_df = final_df[columns_list]

    return final_df

def log_differences(action, sheet_name, num_differences):
    """
    Log the differences found to a text file.
    
    Args:
    action (str): The action taken (e.g., "Added" or "No differences found").
    sheet_name (str): The name of the sheet.
    num_differences (int): The number of differences found.
    """
    time_today = datetime.now().strftime("%Y-%m-%d")
    if num_differences == 0:
        log_message = f"{time_today}: No new data for company: '{sheet_name}'"
    elif num_differences == 1:
        log_message = f"{time_today}: {action} {num_differences} new data row for company: '{sheet_name}'"
    else:
        log_message = f"{time_today}: {action} {num_differences} new data rows for company: '{sheet_name}'"
    with open('log_diff.txt', 'a') as log_file:
        log_file.write(log_message + '\n')

def log_error(message):
    with open("log.txt", "a") as log_file:
        log_file.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {message}\n")

def log_new_and_modified_rows(df1, df2, sheet_name):
    """
    Find rows that are new or modified in df1 compared to df2 and handle Excel file updates.
    
    Args:
    df1 (pd.DataFrame): The most up-to-date dataframe.
    df2 (pd.DataFrame): The older dataframe to compare against.
    sheet_name (str): The name of the sheet to update or delete.
    """
    # Convert columns to the same data type
    for column in df1.columns:
        df1[column] = df1[column].astype(str)

    for column in df2.columns:
        df2[column] = df2[column].astype(str)

    df1['Date of Collection'] = pd.to_datetime(df1['Date of Collection'])
    df2['Date of Collection'] = pd.to_datetime(df2['Date of Collection'])


    # First, we reset the index to ensure the comparison is done row-wise
    df1_reset = df1.reset_index(drop=True)
    df2_reset = df2.reset_index(drop=True)

    excel_file = "Kubrick MI Diff.xlsx"

    # Merge to find new and modified rows
    merged_df = df1_reset.merge(df2_reset, how='left', indicator=True)
    
    # Select new rows
    new_rows = merged_df[merged_df['_merge'] == 'left_only'].drop(columns='_merge')

    # Select potentially modified rows (those that are in both dataframes)
    common_rows = merged_df[merged_df['_merge'] == 'both'].drop(columns='_merge')

    # Identify rows that have changes by comparing the content
    modified_mask = (df1_reset.loc[common_rows.index] != df2_reset.loc[common_rows.index]).any(axis=1)
    modified_rows = df1_reset.loc[common_rows.index][modified_mask]

    # Combine new and modified rows
    new_and_modified_df = pd.concat([new_rows, modified_rows])

    if not os.path.exists(excel_file):
        # Create an empty Excel file with just the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet 1", index=False)

    if not new_and_modified_df.empty:
        # If there are new or modified rows, write them to the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            new_and_modified_df.to_excel(writer, sheet_name=sheet_name, index=False)
        log_differences("Added", sheet_name, len(new_and_modified_df))
    else:
        # If there are no new or modified rows, remove the sheet if it exists
        try:
            workbook = openpyxl.load_workbook(excel_file)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
                workbook.save(excel_file)
        except FileNotFoundError:
            # If the file doesn't exist, do nothing
            pass
        log_differences("No differences found", sheet_name, 0)

def log_new_and_modified_rows2(final_df, old_df, sheet_name, excel_file):
    """
    Find rows that are new or modified in df1 compared to df2 and handle Excel file updates.
    
    Args:
    df1 (pd.DataFrame): The most up-to-date dataframe.
    df2 (pd.DataFrame): The older dataframe to compare against.
    sheet_name (str): The name of the sheet to update or delete.
    """
    # Convert columns to the same data type
    df1 = final_df.copy()
    df2 = old_df.copy()

    cols = list(df1.columns)
    for column in cols:
        df1[column] = df1[column].astype(str)

    df1['Status'] = 'Added'

    for column in cols:
        df2[column] = df2[column].astype(str)

    df2['Status'] = 'Removed'

    new_and_modified_df = pd.concat([df1, df2]).reset_index(drop=True)
    new_and_modified_df.set_index('Status', inplace=True)
    new_and_modified_df['Date of Collection'] = pd.to_datetime(new_and_modified_df['Date of Collection'])
    #df = pd.DataFrame(new_df.groupby(cols).value_counts())

    new_and_modified_df.drop_duplicates(subset=cols, keep=False,inplace=True)
    new_and_modified_df.reset_index(inplace=True)



    if not os.path.exists(excel_file):
        # Create an empty Excel file with just the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
            pd.DataFrame().to_excel(writer, sheet_name="Sheet 1", index=False)

    if not new_and_modified_df.empty:
        # If there are new or modified rows, write them to the specified sheet
        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            new_and_modified_df.to_excel(writer, sheet_name=sheet_name, index=False)
        log_differences("Changed", sheet_name, len(new_and_modified_df))
    else:
        # If there are no new or modified rows, remove the sheet if it exists
        try:
            workbook = openpyxl.load_workbook(excel_file)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
                workbook.save(excel_file)
        except FileNotFoundError:
            # If the file doesn't exist, do nothing
            pass
        log_differences("No differences found", sheet_name, 0)

def get_company_info_pricereport(company_name, ticker, file_path):
    """
    Produces price report for a given company
    
    Args:
    company_name (str): Full company name
    ticker (str): Company ticker. This can be an empty string if the company is private.

    Return:
    (json) : Financial Data to produce the price report
    """
    sheet_name = "Price Report"

    if ticker == "":
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "52 Week Range": "Private"
        }
    
    try:
        # Fetch company data from Yahoo Finance
        stock = yf.Ticker(ticker)
        info = stock.info
        # Check if the returned info indicates data not found
        if 'trailingPegRatio' in info and info['trailingPegRatio'] is None:
            return {
                "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
                "Company Name": company_name,
                "Previous Close": "Private",
                "%-Change": "Private",
                "Sector": "Private",
                "52 Week Range": "Private"
            }
        # Try to get historical data
        hist = stock.history(period="1mo")
        
        if not hist.empty:
            last_close = hist['Close'].iloc[-2]
            prev_close = hist['Close'].iloc[-1]
            percent_change = round(((prev_close - last_close) / last_close) * 100, 4)
        else:
            prev_close = info.get("previousClose", np.nan)
            # Open the Excel file and check for previous close
            wb = openpyxl.load_workbook(file_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                last_close = None
                for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                    if row[1].value == company_name and isinstance(row[2].value, (int, float)):
                        last_close = row[2].value
                        break
                if last_close is None:
                    percent_change = 0
                else:
                    percent_change = round(((prev_close - last_close) / last_close) * 100, 4)
            else:
                percent_change = 0
            wb.close()

        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": info.get("previousClose", np.nan),
            "%-Change": percent_change,
            "Sector": info.get("sector", np.nan),
            "52 Week Range": f'{info.get("fiftyTwoWeekLow", np.nan)} - {info.get("fiftyTwoWeekHigh", np.nan)}'
        }
    except:
        print(f"Error while retrieving financial data for: {ticker}.")
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "52 Week Range": "Private"
        }

def get_company_info_pricereport2(company_name, ticker, file_path):
    """
    Produces price report for a given company
    
    Args:
    company_name (str): Full company name
    ticker (str): Company ticker. This can be an empty string if the company is private.

    Return:
    (json) : Financial Data to produce the price report
    """
    sheet_name = "Price Report"

    if ticker == "":
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "Currency": "Private",
            "52 Week Range": "Private",
            "Error" : False
        }
    
    try:
        # Fetch company data from Yahoo Finance
        stock = yf.Ticker(ticker)
        info = stock.info
        # Check if the returned info indicates data not found
        if len(info) <= 1:
            error_message = f"Issue with ticker: {company_name} : {info}"
            log_error(error_message)
            return {
                "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
                "Company Name": company_name,
                "Previous Close": "Private",
                "%-Change": "Private",
                "Sector": "Private",
                "Currency": "Private",
                "52 Week Range": "Private",
                "Error" : True
                
            }
        # Try to get historical data
        previous_close = info.get("previousClose", "N/A")
        if previous_close == "N/A":
            error_message = f"Error while retrieving financial data for: {ticker}."
            log_error(error_message)
            return {
                "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
                "Company Name": company_name,
                "Previous Close": "Private",
                "%-Change": "Private",
                "Sector": "Private",
                "Currency": "Private",
                "52 Week Range": "Private",
                "Error" : True
            }
        

        # Open the Excel file and check for previous close
        wb = openpyxl.load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            last_close = None
            for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
                if row[1].value == company_name and isinstance(row[2].value, (int, float)):
                    last_close = row[2].value
                    break
            if last_close is None:
                percent_change = 0
            else:
                percent_change = abs(round(((previous_close - last_close) / last_close) * 100, 4))
        else:
            percent_change = 0
        wb.close()
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": previous_close,
            "%-Change": percent_change,
            "Sector": info.get("sector", "No Sector Available"),
            "Currency": info.get('currency', "No Currency Available"),
            "52 Week Range": f'{info.get("fiftyTwoWeekLow", "No 52 Week Low Available")} - {info.get("fiftyTwoWeekHigh", "No 52 Week High Available")}',
            "Error" : False
        }
    except Exception as e:
        error_message = f"Error while retrieving financial data for: {company_name}. Exception: {str(e)}"
        log_error(error_message)
        return {
            "Date of Collection": datetime.now().strftime("%Y-%m-%d"),
            "Company Name": company_name,
            "Previous Close": "Private",
            "%-Change": "Private",
            "Sector": "Private",
            "Currency": "Private",
            "52 Week Range": "Private",
            "Error" : True
        }

def update_excel(data, file_path):

    """
    Updates the sheet "Price Report" with the financial data from a company
    
    Args:
    data (json) : Financial Company data
    """

    sheet_name = "Price Report"

    try:
        # Load workbook and check for sheet
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            # Add headers
            headers = list(data.keys())
            ws.append(headers)
        else:
            ws = wb[sheet_name]
        
        # Find if the company already exists and remove the row if it does
        company_col = 2  # Assuming "Company Name" is in column B (index 1)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
            if row[company_col - 1].value == data["Company Name"]:
                ws.delete_rows(row[0].row)
                break
        
        # Append new data
        ws.append(list(data.values()))

        # Save workbook
        wb.save(file_path)
    except Exception as e:
        print(f"Error updating Excel file: {e}")

def get_scraped_company_data(scraper):
    function_name = scraper
    func = getattr(scrapers, function_name)
    return func()

def column_cleaner(df):
    for column in df.columns:
        column_clean = column.replace('_', ' ')
        df[column] = [f'No {column_clean} Data Available' if data in ('', ' ', None, 'nan', np.nan) else data for data in df[column]]
    
    return df

def table_sorter(df):
    return df.sort_values(by=df.columns.tolist()).reset_index(drop=True)

def create_final_df2(company_name, url, status, json_data, df, template_file_path):
    """
    Creates the final dataframe to be stored using data from webscrape + financial data
    
    Args:
    company_name (str) : Name of company
    url (str) : URL of company
    json_data (json) : Financial data of company
    df (pd.Dataframe) : Webscrape data of comapny

    Returns:
    final_df (pd.Dataframe) : Dataframe with all data in the Template table format
    """
    # Get the current date, company name, and website url metadata
    meta_dict = {'Date of Collection':date.today(), 'Company Name':company_name, 'Website_URL':url, 'Public/Private':status}

    df_copy = df

    columns_list = read_template(template_file_path).columns.tolist()
    for column in columns_list:
        if column in df_copy.columns:
            pass
        elif column in meta_dict.keys():
            df_copy[column] = meta_dict[column]
        elif column in json_data.keys():
            df_copy[column] = json_data[column]
        else:
            df_copy[column] = ' '

    df_clean = column_cleaner(df_copy)
    final_df = df_clean[columns_list]
    final_df_sorted = table_sorter(final_df)

    return final_df_sorted

def get_company_details2(status, symbol):
    """
    Obtains yfinance data if company is public.
    
    Args:
    symbol (str) : Company symbol to obtain yfinance data. This could be blank if company doesn't have a symbol

    Returns:
    (json) : Company information
    """

    if status == "Private":
        company_details = {
                "Previous Close": "Private",
                "52 Week Range": "Private",
                "Sector": "Private",
                "Industry": "Private",
                "Full Time Employees": "Private",
                "Market Cap": "Private",
                "Fiscal Year Ends": "Private",
                "Revenue": "Private",
                "EBITDA": "Private",
        }
    elif status == "Public":
        try:
            stock = yf.Ticker(symbol)
            info = stock.info
            company_details = {
                "Previous Close": info.get("previousClose", np.nan),
                "52 Week Range": f'{info.get("fiftyTwoWeekLow", np.nan)} - {info.get("fiftyTwoWeekHigh", np.nan)}',
                "Sector": info.get("sector", np.nan),
                "Industry": info.get("industry", np.nan),
                "Full Time Employees": info.get("fullTimeEmployees", np.nan),
                "Market Cap": info.get("marketCap", np.nan),
                "Fiscal Year Ends": info.get("fiscalYearEnd", np.nan),
                "Revenue": info.get("totalRevenue", np.nan),
                "EBITDA": info.get("ebitda", np.nan),
            }
        except Exception as e:
            return f'Error, check {symbol} is a valid ticker'
    return company_details

def get_company_metadata(company_name, company_df):
    company_slice = company_df[(company_df['company_name']==company_name)]
    company_url = company_slice.iloc[0]['company_url']
    scraper = company_slice.iloc[0]['scraper']
    status = company_slice.iloc[0]['status']
    ticker = company_slice.iloc[0]['ticker']

    return company_url, scraper, status, ticker